"""
Excel Writer Module Extension for Trade Support
Adds methods to handle trade-related sheets
"""

from typing import Dict, List, Optional
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import pandas as pd
from datetime import datetime
import logging

logger = logging.getLogger(__name__)

# This extends the existing ExcelWriter class
# Add these methods to your existing excel_writer.py

def create_report_with_trades(self, original_positions: List, trade_positions: List, 
                             final_positions: List, prices: Dict[str, float], 
                             unmapped_symbols: List[Dict], unmapped_trades: List[Dict],
                             trades_df: pd.DataFrame = None):
    """Create complete Excel report with trades included"""
    
    # Write original position sheets with "Original_" prefix
    self.write_master_sheet(original_positions, prices, sheet_name="Original_All_Positions")
    
    # Write trade summary sheet
    if trades_df is not None and not trades_df.empty:
        self.write_trades_summary(trades_df)
    
    # Write net trades as positions
    if trade_positions:
        self.write_all_trades_sheet(trade_positions)
    
    # Write final combined positions
    self.write_all_positions_sheet(final_positions, sheet_name="Final_All_Positions")
    
    # Write trade impact analysis
    self.write_trade_impact_sheet(original_positions, trade_positions, final_positions)
    
    # Write final deliverable sheets
    self.write_master_sheet(final_positions, prices, sheet_name="Final_Master_All_Expiries")
    
    # Write final expiry-wise deliverable sheets
    expiries = list(set(p.expiry_date for p in final_positions))
    for expiry in sorted(expiries):
        sheet_name = f"Final_Expiry_{expiry.strftime('%Y_%m_%d')}"
        self.write_expiry_sheet(expiry, final_positions, prices, sheet_name=sheet_name)
    
    # Write final IV sheets
    self.write_iv_master_sheet(final_positions, prices, sheet_name="Final_IV_All_Expiries")
    
    # Write final expiry-wise IV sheets
    for expiry in sorted(expiries):
        sheet_name = f"Final_IV_Expiry_{expiry.strftime('%Y_%m_%d')}"
        self.write_iv_expiry_sheet(expiry, final_positions, prices, sheet_name=sheet_name)
    
    # Write unmapped symbols from both positions and trades
    all_unmapped = unmapped_symbols + unmapped_trades
    if all_unmapped:
        self.write_unmapped_sheet_extended(unmapped_symbols, unmapped_trades)
    
    # Save file
    self.save()

def write_trades_summary(self, trades_df: pd.DataFrame):
    """Write a summary of all trades"""
    ws = self.wb.create_sheet("Trades_Summary")
    
    # Write headers
    headers = list(trades_df.columns)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = self.header_font
        cell.fill = self.header_fill
        cell.alignment = self.header_alignment
        cell.border = self.border
    
    # Write data
    for row_idx, row in trades_df.iterrows():
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx + 2, column=col_idx, value=value)
            cell.border = self.border
            
            # Format numbers
            if col_idx == 7:  # Quantity column
                cell.number_format = '#,##0'
            elif col_idx == 8:  # Price column
                cell.number_format = '#,##0.00'
    
    # Set column widths
    widths = {
        'A': 20,  # TM Name
        'B': 15,  # Symbol
        'C': 12,  # Expiry
        'D': 10,  # Type
        'E': 10,  # Strike
        'F': 8,   # Side
        'G': 12,  # Quantity
        'H': 12   # Price
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

def write_all_trades_sheet(self, trade_positions: List):
    """Write sheet with net positions from trades"""
    ws = self.wb.create_sheet("All_Trades")
    
    headers = ["Underlying", "Symbol", "Expiry", "Net Position", "Type", "Strike", "Lot Size"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = self.header_font
        cell.fill = self.header_fill
        cell.alignment = self.header_alignment
        cell.border = self.border
    
    sorted_positions = sorted(trade_positions, 
                            key=lambda x: (x.underlying_ticker, x.expiry_date, x.strike_price))
    
    current_row = 2
    for pos in sorted_positions:
        ws.cell(row=current_row, column=1, value=pos.underlying_ticker)
        ws.cell(row=current_row, column=2, value=pos.bloomberg_ticker)
        ws.cell(row=current_row, column=3, value=pos.expiry_date.strftime('%Y-%m-%d'))
        
        # Highlight negative (short) positions
        position_cell = ws.cell(row=current_row, column=4, value=pos.position_lots)
        if pos.position_lots < 0:
            position_cell.font = Font(color="FF0000")  # Red for shorts
        
        ws.cell(row=current_row, column=5, value=pos.security_type)
        
        strike_cell = ws.cell(row=current_row, column=6, 
                            value=pos.strike_price if pos.strike_price > 0 else "")
        if pos.strike_price > 0:
            strike_cell.number_format = self.price_format
        
        ws.cell(row=current_row, column=7, value=pos.lot_size)
        
        for col in range(1, 8):
            ws.cell(row=current_row, column=col).border = self.border
        
        current_row += 1
    
    # Set column widths
    widths = {
        'A': 25, 'B': 30, 'C': 12, 'D': 15, 'E': 8, 'F': 10, 'G': 10
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

def write_trade_impact_sheet(self, original_positions: List, trade_positions: List, 
                            final_positions: List):
    """Write sheet showing impact of trades on positions"""
    ws = self.wb.create_sheet("Trade_Impact")
    
    # Create comparison data
    position_map = {}
    
    # Add original positions
    for pos in original_positions:
        key = (pos.underlying_ticker, pos.symbol, pos.expiry_date.date(), 
               pos.security_type, pos.strike_price)
        position_map[key] = {
            'original': pos.position_lots,
            'trade': 0,
            'final': 0,
            'bloomberg': pos.bloomberg_ticker
        }
    
    # Add trade impacts
    for pos in trade_positions:
        key = (pos.underlying_ticker, pos.symbol, pos.expiry_date.date(),
               pos.security_type, pos.strike_price)
        if key not in position_map:
            position_map[key] = {
                'original': 0,
                'trade': pos.position_lots,
                'final': 0,
                'bloomberg': pos.bloomberg_ticker
            }
        else:
            position_map[key]['trade'] = pos.position_lots
    
    # Add final positions
    for pos in final_positions:
        key = (pos.underlying_ticker, pos.symbol, pos.expiry_date.date(),
               pos.security_type, pos.strike_price)
        if key in position_map:
            position_map[key]['final'] = pos.position_lots
    
    # Write headers
    headers = ["Underlying", "Symbol", "Expiry", "Type", "Strike", 
               "Original Position", "Trade Impact", "Final Position", "Change", "Status"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = self.header_font
        cell.fill = self.header_fill
        cell.alignment = self.header_alignment
        cell.border = self.border
    
    # Write data
    current_row = 2
    for key, data in sorted(position_map.items()):
        underlying, symbol, expiry, sec_type, strike = key
        
        ws.cell(row=current_row, column=1, value=underlying)
        ws.cell(row=current_row, column=2, value=data['bloomberg'])
        ws.cell(row=current_row, column=3, value=expiry.strftime('%Y-%m-%d'))
        ws.cell(row=current_row, column=4, value=sec_type)
        ws.cell(row=current_row, column=5, value=strike if strike > 0 else "")
        
        ws.cell(row=current_row, column=6, value=data['original'])
        ws.cell(row=current_row, column=7, value=data['trade'])
        ws.cell(row=current_row, column=8, value=data['final'])
        
        change = data['final'] - data['original']
        change_cell = ws.cell(row=current_row, column=9, value=change)
        
        # Determine status and color
        if data['original'] == 0 and data['final'] != 0:
            status = "NEW"
            row_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        elif data['final'] == 0 and data['original'] != 0:
            status = "CLOSED"
            row_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
        elif data['original'] > 0 and data['final'] < 0:
            status = "FLIPPED SHORT"
            row_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        elif data['original'] < 0 and data['final'] > 0:
            status = "FLIPPED LONG"
            row_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
        elif change > 0:
            status = "INCREASED"
            row_fill = None
        elif change < 0:
            status = "DECREASED"
            row_fill = None
        else:
            status = "UNCHANGED"
            row_fill = None
        
        ws.cell(row=current_row, column=10, value=status)
        
        # Apply formatting
        for col in range(1, 11):
            cell = ws.cell(row=current_row, column=col)
            cell.border = self.border
            if row_fill and col >= 6:  # Highlight position columns
                cell.fill = row_fill
        
        current_row += 1
    
    # Add summary statistics at the bottom
    current_row += 1
    ws.cell(row=current_row, column=1, value="SUMMARY").font = Font(bold=True, size=12)
    current_row += 1
    
    new_positions = sum(1 for k, v in position_map.items() if v['original'] == 0 and v['final'] != 0)
    closed_positions = sum(1 for k, v in position_map.items() if v['final'] == 0 and v['original'] != 0)
    flipped_positions = sum(1 for k, v in position_map.items() 
                           if (v['original'] > 0 and v['final'] < 0) or (v['original'] < 0 and v['final'] > 0))
    
    summary_data = [
        ("New Positions", new_positions),
        ("Closed Positions", closed_positions),
        ("Flipped Positions", flipped_positions),
        ("Total Changes", new_positions + closed_positions + flipped_positions)
    ]
    
    for label, value in summary_data:
        ws.cell(row=current_row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=current_row, column=2, value=value)
        current_row += 1
    
    # Set column widths
    widths = {
        'A': 25, 'B': 30, 'C': 12, 'D': 8, 'E': 10, 
        'F': 18, 'G': 15, 'H': 15, 'I': 12, 'J': 15
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

def write_unmapped_sheet_extended(self, unmapped_positions: List[Dict], unmapped_trades: List[Dict]):
    """Write sheet with unmapped symbols from both positions and trades"""
    ws = self.wb.create_sheet("Unmapped_Symbols")
    
    # Section 1: Unmapped Positions
    ws.cell(row=1, column=1, value="UNMAPPED POSITIONS").font = Font(bold=True, size=12)
    
    headers = ["Symbol", "Expiry", "Position Lots"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = self.header_font
        cell.fill = self.header_fill
        cell.alignment = self.header_alignment
    
    current_row = 3
    for item in unmapped_positions:
        ws.cell(row=current_row, column=1, value=item['symbol'])
        ws.cell(row=current_row, column=2, value=item['expiry'].strftime('%Y-%m-%d'))
        ws.cell(row=current_row, column=3, value=item['position_lots'])
        current_row += 1
    
    # Section 2: Unmapped Trades
    current_row += 2
    ws.cell(row=current_row, column=1, value="UNMAPPED TRADES").font = Font(bold=True, size=12)
    current_row += 1
    
    headers = ["Symbol", "Expiry", "Type", "Strike", "Side", "Quantity"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = self.header_font
        cell.fill = self.header_fill
        cell.alignment = self.header_alignment
    
    current_row += 1
    for item in unmapped_trades:
        ws.cell(row=current_row, column=1, value=item['symbol'])
        ws.cell(row=current_row, column=2, value=item['expiry'].strftime('%Y-%m-%d'))
        ws.cell(row=current_row, column=3, value=item['type'])
        ws.cell(row=current_row, column=4, value=item.get('strike', ''))
        ws.cell(row=current_row, column=5, value=item['side'])
        ws.cell(row=current_row, column=6, value=item['quantity'])
        current_row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 12