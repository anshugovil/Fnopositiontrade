"""
Position Reconciliation Module - Enhanced Version
Supports reconciliation against both Initial and Final positions
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from typing import List, Dict, Tuple, Optional
import logging

logger = logging.getLogger(__name__)


class PositionReconciliation:
    """Handles position reconciliation between delivery output and external file"""
    
    def __init__(self):
        self.header_font = Font(bold=True, size=11)
        self.header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        self.mismatch_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        self.missing_fill = PatternFill(start_color="FFE5CC", end_color="FFE5CC", fill_type="solid")
        self.extra_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def read_positions_sheet(self, excel_file_path: str, sheet_name: str) -> pd.DataFrame:
        """Read a specific positions sheet from delivery calculator output"""
        try:
            # Read the specified sheet
            df = pd.read_excel(excel_file_path, sheet_name=sheet_name)
            
            # We need columns B (Symbol) and D (Position)
            positions_df = pd.DataFrame({
                'Symbol': df.iloc[:, 1],  # Column B - Bloomberg Ticker
                'Position': df.iloc[:, 3]  # Column D - Position
            })
            
            # Clean up - remove any NaN rows
            positions_df = positions_df.dropna()
            
            # Sort by Symbol
            positions_df = positions_df.sort_values('Symbol', ascending=True)
            
            logger.info(f"Read {len(positions_df)} positions from {sheet_name} sheet")
            return positions_df
            
        except Exception as e:
            logger.error(f"Error reading {sheet_name} sheet: {e}")
            raise
    
    def read_all_positions_sheet(self, excel_file_path: str) -> pd.DataFrame:
        """Read All_Positions sheet from delivery calculator output (backward compatibility)"""
        return self.read_positions_sheet(excel_file_path, 'All_Positions')
    
    def read_recon_file(self, recon_file_path: str) -> pd.DataFrame:
        """Read reconciliation file (Excel or CSV)"""
        try:
            # Determine file type and read accordingly
            if recon_file_path.endswith('.csv'):
                df = pd.read_csv(recon_file_path)
            else:
                # Excel file - try to read first sheet
                df = pd.read_excel(recon_file_path)
            
            # Expect columns A (Symbol) and B (Position)
            # Get first two columns regardless of header names
            if df.shape[1] < 2:
                raise ValueError("Recon file must have at least 2 columns")
            
            recon_df = pd.DataFrame({
                'Symbol': df.iloc[:, 0],  # First column
                'Position': df.iloc[:, 1]  # Second column
            })
            
            # Clean up
            recon_df = recon_df.dropna()
            
            # Sort by Symbol
            recon_df = recon_df.sort_values('Symbol', ascending=True)
            
            logger.info(f"Read {len(recon_df)} positions from recon file")
            return recon_df
            
        except Exception as e:
            logger.error(f"Error reading recon file: {e}")
            raise
    
    def reconcile_positions(self, delivery_df: pd.DataFrame, recon_df: pd.DataFrame) -> Dict:
        """
        Compare positions and identify differences
        Returns dictionary with reconciliation results
        """
        # Convert Symbol columns to string for consistent comparison
        delivery_df['Symbol'] = delivery_df['Symbol'].astype(str).str.strip()
        recon_df['Symbol'] = recon_df['Symbol'].astype(str).str.strip()
        
        # Convert Position columns to float for numerical comparison
        delivery_df['Position'] = pd.to_numeric(delivery_df['Position'], errors='coerce')
        recon_df['Position'] = pd.to_numeric(recon_df['Position'], errors='coerce')
        
        # Merge on Symbol to find matches and differences
        merged = pd.merge(
            delivery_df,
            recon_df,
            on='Symbol',
            how='outer',
            suffixes=('_Delivery', '_Recon'),
            indicator=True
        )
        
        # Identify different types of discrepancies
        results = {
            'matched_positions': [],
            'position_mismatches': [],
            'missing_in_recon': [],
            'missing_in_delivery': []
        }
        
        for _, row in merged.iterrows():
            symbol = row['Symbol']
            pos_delivery = row.get('Position_Delivery', 0)
            pos_recon = row.get('Position_Recon', 0)
            
            # Handle NaN values
            pos_delivery = 0 if pd.isna(pos_delivery) else pos_delivery
            pos_recon = 0 if pd.isna(pos_recon) else pos_recon
            
            if row['_merge'] == 'both':
                # Symbol exists in both files
                if abs(pos_delivery - pos_recon) < 0.0001:  # Consider floating point precision
                    results['matched_positions'].append({
                        'Symbol': symbol,
                        'Position': pos_delivery
                    })
                else:
                    results['position_mismatches'].append({
                        'Symbol': symbol,
                        'Delivery_Position': pos_delivery,
                        'Recon_Position': pos_recon,
                        'Difference': pos_delivery - pos_recon
                    })
            elif row['_merge'] == 'left_only':
                # In delivery but not in recon
                results['missing_in_recon'].append({
                    'Symbol': symbol,
                    'Delivery_Position': pos_delivery
                })
            else:  # right_only
                # In recon but not in delivery
                results['missing_in_delivery'].append({
                    'Symbol': symbol,
                    'Recon_Position': pos_recon
                })
        
        # Calculate summary statistics
        results['summary'] = {
            'total_delivery_positions': len(delivery_df),
            'total_recon_positions': len(recon_df),
            'matched_count': len(results['matched_positions']),
            'mismatch_count': len(results['position_mismatches']),
            'missing_in_recon_count': len(results['missing_in_recon']),
            'missing_in_delivery_count': len(results['missing_in_delivery']),
            'total_discrepancies': (
                len(results['position_mismatches']) +
                len(results['missing_in_recon']) +
                len(results['missing_in_delivery'])
            )
        }
        
        return results
    
    def create_dual_recon_report(self, initial_results: Dict, final_results: Dict, output_file: str):
        """Create Excel report with both Initial and Final reconciliation results"""
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # 1. Combined Summary Sheet
        ws_summary = wb.create_sheet("Summary")
        self._write_dual_summary_sheet(ws_summary, initial_results['summary'], final_results['summary'])
        
        # 2. Initial Reconciliation Sheets
        if initial_results['position_mismatches']:
            ws_initial_mismatches = wb.create_sheet("RECON_Initial_Mismatches")
            self._write_mismatches_sheet(ws_initial_mismatches, initial_results['position_mismatches'])
        
        if initial_results['missing_in_recon']:
            ws_initial_missing_recon = wb.create_sheet("RECON_Initial_Missing_in_Recon")
            self._write_missing_sheet(ws_initial_missing_recon, initial_results['missing_in_recon'], 'Delivery')
        
        if initial_results['missing_in_delivery']:
            ws_initial_missing_delivery = wb.create_sheet("RECON_Initial_Missing_in_Delivery")
            self._write_missing_sheet(ws_initial_missing_delivery, initial_results['missing_in_delivery'], 'Recon')
        
        # 3. Final Reconciliation Sheets
        if final_results['position_mismatches']:
            ws_final_mismatches = wb.create_sheet("RECON_Final_Mismatches")
            self._write_mismatches_sheet(ws_final_mismatches, final_results['position_mismatches'])
        
        if final_results['missing_in_recon']:
            ws_final_missing_recon = wb.create_sheet("RECON_Final_Missing_in_Recon")
            self._write_missing_sheet(ws_final_missing_recon, final_results['missing_in_recon'], 'Delivery')
        
        if final_results['missing_in_delivery']:
            ws_final_missing_delivery = wb.create_sheet("RECON_Final_Missing_in_Delivery")
            self._write_missing_sheet(ws_final_missing_delivery, final_results['missing_in_delivery'], 'Recon')
        
        # Save the workbook
        wb.save(output_file)
        logger.info(f"Dual reconciliation report saved: {output_file}")
    
    def create_recon_report(self, recon_results: Dict, output_file: str):
        """Create Excel report with reconciliation results (single reconciliation)"""
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # 1. Summary Sheet
        ws_summary = wb.create_sheet("RECON_vs_Initial_Summary")
        self._write_summary_sheet(ws_summary, recon_results['summary'])
        
        # 2. Position Mismatches Sheet
        if recon_results['position_mismatches']:
            ws_mismatches = wb.create_sheet("RECON_vs_Initial_Mismatches")
            self._write_mismatches_sheet(ws_mismatches, recon_results['position_mismatches'])
        
        # 3. Missing in Recon Sheet
        if recon_results['missing_in_recon']:
            ws_missing_recon = wb.create_sheet("RECON_vs_Initial_Missing_in_Recon")
            self._write_missing_sheet(ws_missing_recon, recon_results['missing_in_recon'], 'Delivery')
        
        # 4. Missing in Delivery Sheet
        if recon_results['missing_in_delivery']:
            ws_missing_delivery = wb.create_sheet("RECON_vs_Initial_Missing_in_Delivery")
            self._write_missing_sheet(ws_missing_delivery, recon_results['missing_in_delivery'], 'Recon')
        
        # 5. All Matched Positions (optional - for verification)
        if recon_results['matched_positions']:
            ws_matched = wb.create_sheet("RECON_vs_Initial_Matched")
            self._write_matched_sheet(ws_matched, recon_results['matched_positions'])
        
        # Save the workbook
        wb.save(output_file)
        logger.info(f"Reconciliation report saved: {output_file}")
    
    def _write_dual_summary_sheet(self, ws, initial_summary, final_summary):
        """Write combined summary statistics sheet for dual reconciliation"""
        ws.cell(row=1, column=1, value="POSITION RECONCILIATION SUMMARY").font = Font(bold=True, size=14)
        
        # Initial Positions Section
        row = 3
        ws.cell(row=row, column=1, value="INITIAL POSITIONS RECONCILIATION").font = Font(bold=True, size=12)
        ws.cell(row=row, column=1).fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        
        row += 1
        initial_items = [
            ("Total Positions in Initial Delivery", initial_summary['total_delivery_positions']),
            ("Total Positions in Recon File", initial_summary['total_recon_positions']),
            ("", ""),
            ("Matched Positions", initial_summary['matched_count']),
            ("Position Mismatches", initial_summary['mismatch_count']),
            ("Missing in Recon File", initial_summary['missing_in_recon_count']),
            ("Missing in Initial Delivery", initial_summary['missing_in_delivery_count']),
            ("", ""),
            ("Total Discrepancies", initial_summary['total_discrepancies'])
        ]
        
        for label, value in initial_items:
            if label:
                ws.cell(row=row, column=1, value=label).font = Font(bold=True)
                ws.cell(row=row, column=2, value=value)
                
                if label == "Total Discrepancies":
                    if value > 0:
                        ws.cell(row=row, column=2).fill = self.mismatch_fill
                    else:
                        ws.cell(row=row, column=2).fill = PatternFill(
                            start_color="90EE90", end_color="90EE90", fill_type="solid"
                        )
            row += 1
        
        # Final Positions Section
        row += 2
        ws.cell(row=row, column=1, value="FINAL POSITIONS RECONCILIATION (AFTER TRADES)").font = Font(bold=True, size=12)
        ws.cell(row=row, column=1).fill = PatternFill(start_color="FFE5CC", end_color="FFE5CC", fill_type="solid")
        
        row += 1
        final_items = [
            ("Total Positions in Final Delivery", final_summary['total_delivery_positions']),
            ("Total Positions in Recon File", final_summary['total_recon_positions']),
            ("", ""),
            ("Matched Positions", final_summary['matched_count']),
            ("Position Mismatches", final_summary['mismatch_count']),
            ("Missing in Recon File", final_summary['missing_in_recon_count']),
            ("Missing in Final Delivery", final_summary['missing_in_delivery_count']),
            ("", ""),
            ("Total Discrepancies", final_summary['total_discrepancies'])
        ]
        
        for label, value in final_items:
            if label:
                ws.cell(row=row, column=1, value=label).font = Font(bold=True)
                ws.cell(row=row, column=2, value=value)
                
                if label == "Total Discrepancies":
                    if value > 0:
                        ws.cell(row=row, column=2).fill = self.mismatch_fill
                    else:
                        ws.cell(row=row, column=2).fill = PatternFill(
                            start_color="90EE90", end_color="90EE90", fill_type="solid"
                        )
            row += 1
        
        # Add borders
        for r in range(3, row):
            for c in range(1, 3):
                if ws.cell(row=r, column=c).value:
                    ws.cell(row=r, column=c).border = self.border
        
        # Set column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 15
    
    def _write_summary_sheet(self, ws, summary):
        """Write summary statistics sheet"""
        ws.cell(row=1, column=1, value="POSITION RECONCILIATION SUMMARY").font = Font(bold=True, size=14)
        
        row = 3
        summary_items = [
            ("Total Positions in Delivery Output", summary['total_delivery_positions']),
            ("Total Positions in Recon File", summary['total_recon_positions']),
            ("", ""),
            ("Matched Positions", summary['matched_count']),
            ("Position Mismatches", summary['mismatch_count']),
            ("Missing in Recon File", summary['missing_in_recon_count']),
            ("Missing in Delivery Output", summary['missing_in_delivery_count']),
            ("", ""),
            ("Total Discrepancies", summary['total_discrepancies'])
        ]
        
        for label, value in summary_items:
            if label:
                ws.cell(row=row, column=1, value=label).font = Font(bold=True)
                ws.cell(row=row, column=2, value=value)
                
                if label == "Total Discrepancies":
                    if value > 0:
                        ws.cell(row=row, column=2).fill = self.mismatch_fill
                    else:
                        ws.cell(row=row, column=2).fill = PatternFill(
                            start_color="90EE90", end_color="90EE90", fill_type="solid"
                        )
            row += 1
        
        # Add borders
        for r in range(3, row):
            for c in range(1, 3):
                ws.cell(row=r, column=c).border = self.border
        
        # Set column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
    
    def _write_mismatches_sheet(self, ws, mismatches):
        """Write position mismatches sheet"""
        headers = ["Symbol", "Delivery Position", "Recon Position", "Difference", "Abs Difference"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        
        row = 2
        for item in sorted(mismatches, key=lambda x: abs(x['Difference']), reverse=True):
            ws.cell(row=row, column=1, value=item['Symbol'])
            ws.cell(row=row, column=2, value=item['Delivery_Position'])
            ws.cell(row=row, column=3, value=item['Recon_Position'])
            ws.cell(row=row, column=4, value=item['Difference'])
            ws.cell(row=row, column=5, value=abs(item['Difference']))
            
            # Highlight the row
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = self.mismatch_fill
                ws.cell(row=row, column=col).border = self.border
            
            row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
    
    def _write_missing_sheet(self, ws, missing_items, source_type):
        """Write missing positions sheet"""
        headers = ["Symbol", f"{source_type} Position"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        
        row = 2
        for item in sorted(missing_items, key=lambda x: x['Symbol']):
            ws.cell(row=row, column=1, value=item['Symbol'])
            
            position_key = f'{source_type}_Position'
            ws.cell(row=row, column=2, value=item[position_key])
            
            # Highlight the row
            fill_color = self.missing_fill if source_type == 'Delivery' else self.extra_fill
            for col in range(1, 3):
                ws.cell(row=row, column=col).fill = fill_color
                ws.cell(row=row, column=col).border = self.border
            
            row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 18
    
    def _write_matched_sheet(self, ws, matched_items):
        """Write matched positions sheet"""
        headers = ["Symbol", "Position"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        
        row = 2
        for item in sorted(matched_items, key=lambda x: x['Symbol']):
            ws.cell(row=row, column=1, value=item['Symbol'])
            ws.cell(row=row, column=2, value=item['Position'])
            
            for col in range(1, 3):
                ws.cell(row=row, column=col).border = self.border
            
            row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 18
    
    def perform_reconciliation(self, delivery_file: str, recon_file: str, output_file: str, 
                             has_trades: bool = False) -> Dict:
        """
        Main method to perform reconciliation
        If has_trades is True, reconciles against both Initial and Final positions
        """
        try:
            recon_df = self.read_recon_file(recon_file)
            
            if has_trades:
                # Dual reconciliation - against both Initial and Final
                initial_df = self.read_positions_sheet(delivery_file, 'Start_All_Positions')
                final_df = self.read_positions_sheet(delivery_file, 'Final_All_Positions')
                
                initial_results = self.reconcile_positions(initial_df, recon_df)
                final_results = self.reconcile_positions(final_df, recon_df)
                
                self.create_dual_recon_report(initial_results, final_results, output_file)
                
                return {
                    'initial': initial_results,
                    'final': final_results
                }
            else:
                # Single reconciliation - against All_Positions
                delivery_df = self.read_all_positions_sheet(delivery_file)
                results = self.reconcile_positions(delivery_df, recon_df)
                self.create_recon_report(results, output_file)
                
                return {
                    'initial': results
                }
            
        except Exception as e:
            logger.error(f"Error during reconciliation: {e}")
            raise
