"""
Streamlit Futures Delivery Calculator with Trade File Support
Web application for calculating physical delivery with optional trade positions
"""

import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import io
import tempfile
import os
import logging
from typing import Dict, List, Optional
from dataclasses import dataclass

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Define Position class here to avoid import issues
@dataclass
class Position:
    """Represents a single position"""
    underlying_ticker: str
    bloomberg_ticker: str
    symbol: str
    expiry_date: datetime
    position_lots: float
    security_type: str  # Futures, Call, Put
    strike_price: float
    lot_size: int
    
    @property
    def is_future(self) -> bool:
        return self.security_type == 'Futures'
    
    @property
    def is_call(self) -> bool:
        return self.security_type == 'Call'
    
    @property
    def is_put(self) -> bool:
        return self.security_type == 'Put'

# Import modules
from input_parser import InputParser
from trade_parser import TradeParser
from price_fetcher import PriceFetcher
from excel_writer import ExcelWriter
from recon_module import PositionReconciliation

# Page config
st.set_page_config(
    page_title="Futures Delivery Calculator",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for better styling
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        color: #1f77b4;
        text-align: center;
        padding: 1rem;
        border-bottom: 3px solid #1f77b4;
        margin-bottom: 2rem;
    }
    .sub-header {
        font-size: 1.5rem;
        color: #333;
        margin-top: 1.5rem;
        margin-bottom: 1rem;
    }
    .info-box {
        background-color: #f0f2f6;
        padding: 1rem;
        border-radius: 0.5rem;
        margin: 1rem 0;
    }
    .success-box {
        background-color: #d4edda;
        padding: 1rem;
        border-radius: 0.5rem;
        margin: 1rem 0;
        border: 1px solid #c3e6cb;
    }
    .trade-box {
        background-color: #fff3cd;
        padding: 1rem;
        border-radius: 0.5rem;
        margin: 1rem 0;
        border: 1px solid #ffc107;
    }
</style>
""", unsafe_allow_html=True)


class StreamlitDeliveryApp:
    """Main Streamlit application class with trade support"""
    
    def __init__(self):
        self.initialize_session_state()
        self.recon_module = PositionReconciliation()
    
    def initialize_session_state(self):
        """Initialize session state variables"""
        if 'positions' not in st.session_state:
            st.session_state.positions = []
        if 'trade_positions' not in st.session_state:
            st.session_state.trade_positions = []
        if 'final_positions' not in st.session_state:
            st.session_state.final_positions = []
        if 'prices' not in st.session_state:
            st.session_state.prices = {}
        if 'unmapped_symbols' not in st.session_state:
            st.session_state.unmapped_symbols = []
        if 'report_generated' not in st.session_state:
            st.session_state.report_generated = False
        if 'output_file' not in st.session_state:
            st.session_state.output_file = None
        if 'recon_results' not in st.session_state:
            st.session_state.recon_results = None
        if 'recon_file' not in st.session_state:
            st.session_state.recon_file = None
        if 'file_prefix' not in st.session_state:
            st.session_state.file_prefix = 'DELIVERY'
        if 'has_trades' not in st.session_state:
            st.session_state.has_trades = False
        if 'trade_file' not in st.session_state:
            st.session_state.trade_file = None
        if 'trade_format' not in st.session_state:
            st.session_state.trade_format = 'Auto-detect'
    
    def run(self):
        """Main application entry point"""
        # Header
        st.markdown('<h1 class="main-header">📊 Futures & Options Delivery Calculator</h1>', 
                   unsafe_allow_html=True)
        
        # Sidebar for configuration
        with st.sidebar:
            st.header("⚙️ Configuration")
            
            # USDINR Rate
            usdinr_rate = st.number_input(
                "USD/INR Exchange Rate",
                min_value=50.0,
                max_value=150.0,
                value=88.0,
                step=0.1,
                help="Current USD to INR exchange rate for IV calculations"
            )
            
            # Mapping file upload
            st.subheader("📁 Symbol Mapping File")
            mapping_file = st.file_uploader(
                "Upload futures mapping CSV",
                type=['csv'],
                help="CSV file with symbol to ticker mappings"
            )
            
            mapping_file_path = None
            if not mapping_file:
                st.info("ℹ️ Using default 'futures mapping.csv'")
                possible_paths = ['futures mapping.csv', 'futures_mapping.csv']
                for path in possible_paths:
                    if os.path.exists(path):
                        mapping_file_path = path
                        break
                if not mapping_file_path:
                    mapping_file_path = 'futures mapping.csv'
            else:
                # Save uploaded mapping file temporarily
                with tempfile.NamedTemporaryFile(delete=False, suffix='.csv', mode='wb') as tmp_file:
                    tmp_file.write(mapping_file.getvalue())
                    mapping_file_path = tmp_file.name
            
            st.divider()
            
            # Trade file upload section
            st.subheader("📈 Trade File (Optional)")
            st.info("Upload a trade file to calculate final positions after trades")
            trade_file = st.file_uploader(
                "Upload MS/GS trade file",
                type=['xlsx', 'xls', 'csv'],
                help="Trade file in MS or GS format",
                key="trade_uploader"
            )
            
            if trade_file:
                st.success(f"✅ Trade file loaded: {trade_file.name}")
                st.session_state.trade_file = trade_file
            
            st.divider()
            
            # Price fetching options
            st.subheader("💹 Price Options")
            fetch_prices = st.checkbox("Fetch prices from Yahoo Finance", value=True)
            
            st.divider()
            
            # Reconciliation options
            st.subheader("🔄 Reconciliation (Optional)")
            st.info("Upload a recon file to compare positions")
            recon_file = st.file_uploader(
                "Upload reconciliation file",
                type=['xlsx', 'xls', 'csv'],
                help="File with Symbol and Position columns to reconcile against",
                key="recon_uploader"
            )
            
            if recon_file:
                st.success(f"✅ Recon file loaded: {recon_file.name}")
                st.session_state.recon_file = recon_file
        
        # Main content area with tabs
        tabs = st.tabs(["📤 Upload & Process", "📊 Positions Review", 
                        "💼 Trade Positions", "📥 Download Reports"])
        
        with tabs[0]:
            self.upload_and_process_tab(mapping_file_path, usdinr_rate, fetch_prices)
        
        with tabs[1]:
            self.positions_review_tab()
        
        with tabs[2]:
            self.trade_positions_tab()
        
        with tabs[3]:
            self.download_reports_tab()
    
    def upload_and_process_tab(self, mapping_file_path, usdinr_rate, fetch_prices):
        """Handle file upload and processing"""
        st.markdown('<h2 class="sub-header">Upload Position File</h2>', unsafe_allow_html=True)
        
        col1, col2 = st.columns([2, 1])
        
        with col1:
            uploaded_file = st.file_uploader(
                "Choose your position file",
                type=['xlsx', 'xls', 'csv'],
                help="Upload BOD, CONTRACT, or MS format position file"
            )
        
        with col2:
            if uploaded_file:
                st.markdown('<div class="info-box">', unsafe_allow_html=True)
                st.write("**File Details:**")
                st.write(f"📁 Name: {uploaded_file.name}")
                st.write(f"📏 Size: {uploaded_file.size:,} bytes")
                st.markdown('</div>', unsafe_allow_html=True)
        
        if uploaded_file and mapping_file_path:
            # Process button
            if st.button("🚀 Process Files", type="primary", use_container_width=True):
                with st.spinner("Processing position file..."):
                    # Process position file
                    success, message = self.process_file(
                        uploaded_file, mapping_file_path, usdinr_rate, fetch_prices
                    )
                    
                    if success:
                        st.success(f"✅ {message}")
                        
                        # Process trade file if uploaded
                        if st.session_state.trade_file:
                            with st.spinner("Processing trade file..."):
                                trade_success = self.process_trade_file(
                                    st.session_state.trade_file, 
                                    mapping_file_path,
                                    fetch_prices
                                )
                                if trade_success:
                                    st.success("✅ Trade file processed successfully")
                                    st.session_state.has_trades = True
                        
                        # Generate the report
                        self.generate_final_report(usdinr_rate)
                        
                        # If recon file is uploaded, perform reconciliation
                        if st.session_state.recon_file:
                            self.perform_reconciliation()
                        
                        st.balloons()
                    else:
                        st.error(f"❌ {message}")
    
    def process_trade_file(self, trade_file, mapping_file_path, fetch_prices):
        """Process the trade file"""
        try:
            # Save trade file temporarily
            suffix = os.path.splitext(trade_file.name)[1]
            with tempfile.NamedTemporaryFile(delete=False, suffix=suffix, mode='wb') as tmp_file:
                tmp_file.write(trade_file.getvalue())
                trade_file_path = tmp_file.name
            
            # Parse trades
            trade_parser = TradeParser(mapping_file_path)
            trade_positions = trade_parser.parse_trade_file(trade_file_path)
            
            if not trade_positions:
                st.warning("No valid trade positions found in trade file")
                return False
            
            st.session_state.trade_positions = trade_positions
            
            # Fetch additional prices if needed
            if fetch_prices:
                from price_fetcher import PriceFetcher
                symbols_to_fetch = list(set(p.symbol for p in trade_positions))
                if symbols_to_fetch:
                    price_fetcher = PriceFetcher()
                    symbol_prices = price_fetcher.fetch_prices_for_symbols(symbols_to_fetch)
                    
                    for pos in trade_positions:
                        if pos.symbol in symbol_prices:
                            if pos.underlying_ticker not in st.session_state.prices:
                                st.session_state.prices[pos.underlying_ticker] = symbol_prices[pos.symbol]
            
            # Calculate final positions
            self.calculate_final_positions()
            
            # Clean up temp file
            try:
                os.unlink(trade_file_path)
            except:
                pass
            
            return True
            
        except Exception as e:
            logger.error(f"Error processing trade file: {str(e)}")
            st.error(f"Error processing trade file: {str(e)}")
            import traceback
            st.code(traceback.format_exc())
            return False
    
    def calculate_final_positions(self):
        """Calculate final positions by combining start and trade positions"""
        position_map = {}
        
        # Helper function to create position key
        def make_key(pos):
            return (pos.underlying_ticker, pos.bloomberg_ticker, pos.symbol,
                   pos.expiry_date, pos.security_type, pos.strike_price)
        
        # Add start positions
        for pos in st.session_state.positions:
            key = make_key(pos)
            if key not in position_map:
                position_map[key] = {
                    'position': pos,
                    'net_lots': pos.position_lots
                }
            else:
                position_map[key]['net_lots'] += pos.position_lots
        
        # Add trade positions
        for pos in st.session_state.trade_positions:
            key = make_key(pos)
            if key not in position_map:
                position_map[key] = {
                    'position': pos,
                    'net_lots': pos.position_lots
                }
            else:
                position_map[key]['net_lots'] += pos.position_lots
        
        # Create final positions list
        final_positions = []
        for key, data in position_map.items():
            pos = data['position']
            final_positions.append(Position(
                underlying_ticker=pos.underlying_ticker,
                bloomberg_ticker=pos.bloomberg_ticker,
                symbol=pos.symbol,
                expiry_date=pos.expiry_date,
                position_lots=data['net_lots'],
                security_type=pos.security_type,
                strike_price=pos.strike_price,
                lot_size=pos.lot_size
            ))
        
        st.session_state.final_positions = final_positions
    
    def generate_final_report(self, usdinr_rate):
        """Generate the final Excel report"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            prefix = st.session_state.file_prefix
            
            if st.session_state.has_trades:
                output_file = f"{prefix}_WITH_TRADES_{timestamp}.xlsx"
            else:
                output_file = f"{prefix}_{timestamp}.xlsx"
            
            writer = ExcelWriter(output_file, usdinr_rate)
            
            # Pass trade positions if available
            trade_positions = st.session_state.trade_positions if st.session_state.has_trades else None
            
            writer.create_report(
                st.session_state.positions,
                st.session_state.prices,
                st.session_state.unmapped_symbols,
                trade_positions
            )
            
            st.session_state.output_file = output_file
            st.session_state.report_generated = True
            
        except Exception as e:
            logger.error(f"Error generating report: {str(e)}")
            st.error(f"Error generating report: {str(e)}")
    
    def process_file(self, uploaded_file, mapping_file_path, usdinr_rate, fetch_prices):
        """Process the uploaded position file"""
        try:
            # Save uploaded file temporarily
            suffix = os.path.splitext(uploaded_file.name)[1]
            with tempfile.NamedTemporaryFile(delete=False, suffix=suffix, mode='wb') as tmp_file:
                tmp_file.write(uploaded_file.getvalue())
                input_file_path = tmp_file.name
            
            # Parse positions
            parser = InputParser(mapping_file_path)
            positions = parser.parse_file(input_file_path)
            
            if not positions:
                return False, "No valid positions found in the file"
            
            st.session_state.positions = positions
            st.session_state.unmapped_symbols = parser.unmapped_symbols
            
            # Determine file prefix based on format
            format_type = getattr(parser, 'format_type', 'UNKNOWN')
            if format_type in ['BOD', 'CONTRACT']:
                st.session_state.file_prefix = "GS_AURIGIN"
            elif format_type == 'MS':
                st.session_state.file_prefix = "MS_WAFRA"
            else:
                st.session_state.file_prefix = "DELIVERY"
            
            # Fetch prices if enabled
            if fetch_prices:
                with st.spinner("Fetching prices from Yahoo Finance..."):
                    from price_fetcher import PriceFetcher
                    price_fetcher = PriceFetcher()
                    symbols_to_fetch = list(set(p.symbol for p in positions))
                    symbol_prices = price_fetcher.fetch_prices_for_symbols(symbols_to_fetch)
                    
                    # Map to underlying tickers
                    symbol_map = {}
                    for p in positions:
                        symbol_map[p.underlying_ticker] = p.symbol
                    
                    prices = {}
                    for underlying, symbol in symbol_map.items():
                        if symbol in symbol_prices:
                            prices[underlying] = symbol_prices[symbol]
                    
                    st.session_state.prices = prices
            
            # Clean up temp file
            try:
                os.unlink(input_file_path)
            except:
                pass
            
            return True, f"Successfully processed {len(positions)} positions"
            
        except Exception as e:
            logger.error(f"Error processing file: {str(e)}")
            return False, f"Error processing file: {str(e)}"
    
    def perform_reconciliation(self):
        """Perform reconciliation"""
        if not st.session_state.output_file or not st.session_state.recon_file:
            return
        
        try:
            # Save recon file temporarily
            suffix = os.path.splitext(st.session_state.recon_file.name)[1]
            with tempfile.NamedTemporaryFile(delete=False, suffix=suffix, mode='wb') as tmp_file:
                tmp_file.write(st.session_state.recon_file.getvalue())
                recon_file_path = tmp_file.name
            
            # Generate recon output filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            prefix = st.session_state.file_prefix
            recon_output_file = f"{prefix}_RECONCILIATION_{timestamp}.xlsx"
            
            # Perform reconciliation (dual if trades, single if no trades)
            results = self.recon_module.perform_reconciliation(
                st.session_state.output_file,
                recon_file_path,
                recon_output_file,
                has_trades=st.session_state.has_trades
            )
            
            st.session_state.recon_results = results
            st.session_state.recon_output_file = recon_output_file
            
            # Clean up temp file
            try:
                os.unlink(recon_file_path)
            except:
                pass
                
        except Exception as e:
            logger.error(f"Error during reconciliation: {str(e)}")
            st.error(f"Reconciliation failed: {str(e)}")
    
    def trade_positions_tab(self):
        """Display trade positions"""
        st.markdown('<h2 class="sub-header">Trade Positions</h2>', unsafe_allow_html=True)
        
        if not st.session_state.has_trades:
            st.info("📤 No trade file has been processed yet")
            return
        
        trade_positions = st.session_state.trade_positions
        
        # Summary metrics
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric("Total Trade Lines", len(trade_positions))
        
        with col2:
            buys = sum(1 for p in trade_positions if p.position_lots > 0)
            st.metric("Buy Trades", buys)
        
        with col3:
            sells = sum(1 for p in trade_positions if p.position_lots < 0)
            st.metric("Sell Trades", sells)
        
        with col4:
            net_lots = sum(p.position_lots for p in trade_positions)
            st.metric("Net Position", f"{net_lots:+.0f}")
        
        # Trade positions table
        st.subheader("📋 Trade Details")
        
        df_data = []
        for p in trade_positions:
            df_data.append({
                'Underlying': p.underlying_ticker,
                'Symbol': p.symbol,
                'Bloomberg Ticker': p.bloomberg_ticker,
                'Expiry': p.expiry_date.strftime('%Y-%m-%d'),
                'Type': p.security_type,
                'Strike': p.strike_price if p.strike_price > 0 else '',
                'Position (Lots)': p.position_lots,
                'Side': 'BUY' if p.position_lots > 0 else 'SELL',
                'Lot Size': p.lot_size
            })
        
        if df_data:
            df = pd.DataFrame(df_data)
            st.dataframe(df, use_container_width=True, hide_index=True)
    
    def positions_review_tab(self):
        """Display parsed positions for review"""
        st.markdown('<h2 class="sub-header">Position Summary</h2>', unsafe_allow_html=True)
        
        if not st.session_state.positions:
            st.info("📤 Please upload and process a position file first")
            return
        
        # If we have trades, show both initial and final positions
        if st.session_state.has_trades and st.session_state.final_positions:
            position_view = st.radio(
                "Select positions to view:",
                ["Initial Positions", "Final Positions (After Trades)", "Compare Both"]
            )
            
            if position_view == "Initial Positions":
                self._display_positions_section("Initial Positions", st.session_state.positions)
            elif position_view == "Final Positions (After Trades)":
                self._display_positions_section("Final Positions (After Trades)", st.session_state.final_positions)
            else:
                # Compare both
                col1, col2 = st.columns(2)
                with col1:
                    st.markdown("### Initial Positions")
                    self._display_positions_metrics(st.session_state.positions)
                with col2:
                    st.markdown("### Final Positions")
                    self._display_positions_metrics(st.session_state.final_positions)
                
                # Show detailed comparison
                st.subheader("📋 Position Comparison")
                comparison_data = self._create_comparison_data(
                    st.session_state.positions, 
                    st.session_state.final_positions
                )
                if comparison_data:
                    comp_df = pd.DataFrame(comparison_data)
                    st.dataframe(comp_df, use_container_width=True, hide_index=True)
        else:
            # No trades, just show initial positions
            self._display_positions_section("Initial Positions", st.session_state.positions)
        
        # Unmapped symbols warning
        if st.session_state.unmapped_symbols:
            st.warning(f"⚠️ {len(st.session_state.unmapped_symbols)} unmapped symbols found")
            with st.expander("View Unmapped Symbols"):
                unmapped_df = pd.DataFrame(st.session_state.unmapped_symbols)
                st.dataframe(unmapped_df, use_container_width=True, hide_index=True)
    
    def _display_positions_section(self, title, positions):
        """Display a positions section with metrics and details"""
        # Summary metrics
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric("Total Positions", len(positions))
        
        with col2:
            unique_underlyings = len(set(p.underlying_ticker for p in positions))
            st.metric("Unique Underlyings", unique_underlyings)
        
        with col3:
            unique_expiries = len(set(p.expiry_date for p in positions))
            st.metric("Unique Expiries", unique_expiries)
        
        with col4:
            futures_count = sum(1 for p in positions if p.is_future)
            options_count = len(positions) - futures_count
            st.metric("Futures/Options", f"{futures_count}/{options_count}")
        
        # Detailed positions table
        st.subheader(f"📋 {title} Details")
        
        df_data = []
        for p in positions:
            df_data.append({
                'Underlying': p.underlying_ticker,
                'Symbol': p.symbol,
                'Bloomberg Ticker': p.bloomberg_ticker,
                'Expiry': p.expiry_date.strftime('%Y-%m-%d'),
                'Type': p.security_type,
                'Strike': p.strike_price if p.strike_price > 0 else '',
                'Position (Lots)': p.position_lots,
                'Lot Size': p.lot_size
            })
        
        if df_data:
            df = pd.DataFrame(df_data)
            # Highlight closed positions (0 lots) if this is final positions
            if "Final" in title:
                styled_df = df.style.apply(
                    lambda x: ['background-color: #ffe6e6' if x['Position (Lots)'] == 0 else '' for _ in x],
                    axis=1
                )
                st.dataframe(df, use_container_width=True, hide_index=True)
            else:
                st.dataframe(df, use_container_width=True, hide_index=True)
    
    def _display_positions_metrics(self, positions):
        """Display just the metrics for positions"""
        st.metric("Total", len(positions))
        st.metric("Underlyings", len(set(p.underlying_ticker for p in positions)))
        st.metric("Expiries", len(set(p.expiry_date for p in positions)))
        futures = sum(1 for p in positions if p.is_future)
        st.metric("Fut/Opt", f"{futures}/{len(positions)-futures}")
    
    def _create_comparison_data(self, initial_positions, final_positions):
        """Create comparison data between initial and final positions"""
        # Create a map of positions
        position_map = {}
        
        # Add initial positions
        for pos in initial_positions:
            key = (pos.underlying_ticker, pos.bloomberg_ticker, pos.expiry_date, pos.security_type, pos.strike_price)
            position_map[key] = {
                'Underlying': pos.underlying_ticker,
                'Symbol': pos.symbol,
                'Expiry': pos.expiry_date.strftime('%Y-%m-%d'),
                'Type': pos.security_type,
                'Strike': pos.strike_price if pos.strike_price > 0 else 0,  # Use 0 instead of empty string
                'Initial': pos.position_lots,
                'Final': 0,
                'Change': 0
            }
        
        # Add final positions
        for pos in final_positions:
            key = (pos.underlying_ticker, pos.bloomberg_ticker, pos.expiry_date, pos.security_type, pos.strike_price)
            if key in position_map:
                position_map[key]['Final'] = pos.position_lots
                position_map[key]['Change'] = pos.position_lots - position_map[key]['Initial']
            else:
                position_map[key] = {
                    'Underlying': pos.underlying_ticker,
                    'Symbol': pos.symbol,
                    'Expiry': pos.expiry_date.strftime('%Y-%m-%d'),
                    'Type': pos.security_type,
                    'Strike': pos.strike_price if pos.strike_price > 0 else 0,  # Use 0 instead of empty string
                    'Initial': 0,
                    'Final': pos.position_lots,
                    'Change': pos.position_lots
                }
        
        # Convert to list and sort
        comparison_data = list(position_map.values())
        # Sort with proper handling of Strike field
        comparison_data.sort(key=lambda x: (x['Underlying'], x['Expiry'], float(x['Strike']) if x['Strike'] else 0))
        
        # Convert Strike back to display format (empty string for 0)
        for item in comparison_data:
            if item['Strike'] == 0:
                item['Strike'] = ''
        
        return comparison_data
    
    def reconciliation_tab(self):
        """Display reconciliation results"""
        st.markdown('<h2 class="sub-header">Position Reconciliation</h2>', unsafe_allow_html=True)
        
        if not st.session_state.report_generated:
            st.info("📤 Please process a position file first")
            return
        
        if not st.session_state.recon_file:
            st.info("📋 Upload a reconciliation file in the sidebar to compare positions")
            st.write("The recon file should have two columns:")
            st.write("- Column A: Symbol (Bloomberg Ticker)")
            st.write("- Column B: Position")
            return
        
        # Run reconciliation button
        if not st.session_state.recon_results:
            if st.button("🔄 Run Reconciliation", type="primary", use_container_width=True):
                with st.spinner("Performing reconciliation..."):
                    self.perform_reconciliation()
                    st.success("✅ Reconciliation completed!")
                    st.rerun()  # Rerun to show results
            return
        
        # Display reconciliation results
        if st.session_state.recon_results:
            results = st.session_state.recon_results
            
            # Check if we have dual reconciliation (with trades)
            if st.session_state.has_trades and 'final' in results:
                # Dual reconciliation display
                st.subheader("📊 Dual Reconciliation Summary")
                st.info("Reconciliation performed against both Initial and Final positions")
                
                # Summary metrics in columns
                col1, col2 = st.columns(2)
                
                with col1:
                    st.markdown("### Initial Positions Reconciliation")
                    initial_summary = results['initial']['summary']
                    
                    # Metrics
                    mcol1, mcol2 = st.columns(2)
                    with mcol1:
                        st.metric("Matched", initial_summary['matched_count'])
                        st.metric("Mismatches", initial_summary['mismatch_count'])
                    with mcol2:
                        st.metric("Missing in Recon", initial_summary['missing_in_recon_count'])
                        st.metric("Missing in Delivery", initial_summary['missing_in_delivery_count'])
                    
                    # Total discrepancies
                    if initial_summary['total_discrepancies'] > 0:
                        st.error(f"⚠️ Total Discrepancies: {initial_summary['total_discrepancies']}")
                    else:
                        st.success("✅ All positions match!")
                
                with col2:
                    st.markdown("### Final Positions Reconciliation")
                    final_summary = results['final']['summary']
                    
                    # Metrics
                    mcol1, mcol2 = st.columns(2)
                    with mcol1:
                        st.metric("Matched", final_summary['matched_count'])
                        st.metric("Mismatches", final_summary['mismatch_count'])
                    with mcol2:
                        st.metric("Missing in Recon", final_summary['missing_in_recon_count'])
                        st.metric("Missing in Delivery", final_summary['missing_in_delivery_count'])
                    
                    # Total discrepancies
                    if final_summary['total_discrepancies'] > 0:
                        st.error(f"⚠️ Total Discrepancies: {final_summary['total_discrepancies']}")
                    else:
                        st.success("✅ All positions match!")
                
                st.divider()
                
                # Detailed view selector
                st.subheader("📋 Detailed Reconciliation View")
                recon_view = st.radio(
                    "Select reconciliation details to view:",
                    ["Initial Positions vs Recon", "Final Positions vs Recon", "Compare Both"]
                )
                
                if recon_view == "Compare Both":
                    # Show comparison of both
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        st.markdown("#### Initial Positions Discrepancies")
                        self._display_recon_details(results['initial'], "Initial")
                    
                    with col2:
                        st.markdown("#### Final Positions Discrepancies")
                        self._display_recon_details(results['final'], "Final")
                
                elif recon_view == "Initial Positions vs Recon":
                    self._display_recon_details(results['initial'], "Initial")
                
                else:  # Final Positions vs Recon
                    self._display_recon_details(results['final'], "Final")
            
            else:
                # Single reconciliation display (no trades)
                st.subheader("📊 Reconciliation Summary")
                st.info("Reconciliation performed against Initial positions only (no trades)")
                
                display_results = results['initial']
                summary = display_results['summary']
                
                # Summary metrics
                col1, col2, col3, col4 = st.columns(4)
                
                with col1:
                    st.metric("Matched Positions", summary['matched_count'])
                
                with col2:
                    st.metric("Position Mismatches", summary['mismatch_count'])
                
                with col3:
                    st.metric("Missing in Recon", summary['missing_in_recon_count'])
                
                with col4:
                    st.metric("Missing in Delivery", summary['missing_in_delivery_count'])
                
                # Total discrepancies
                if summary['total_discrepancies'] > 0:
                    st.error(f"⚠️ Total Discrepancies: {summary['total_discrepancies']}")
                else:
                    st.success("✅ All positions match perfectly!")
                
                st.divider()
                
                # Display detailed discrepancies
                st.subheader("📋 Detailed Discrepancies")
                self._display_recon_details(display_results, "Initial")
    
    def _display_recon_details(self, recon_results, label):
        """Display detailed reconciliation discrepancies"""
        # Position mismatches
        if recon_results.get('position_mismatches'):
            st.write(f"**🔍 Position Mismatches ({label})**")
            mismatch_df = pd.DataFrame(recon_results['position_mismatches'])
            st.dataframe(
                mismatch_df,
                use_container_width=True,
                hide_index=True,
                column_config={
                    'Delivery_Position': st.column_config.NumberColumn(format="%.2f"),
                    'Recon_Position': st.column_config.NumberColumn(format="%.2f"),
                    'Difference': st.column_config.NumberColumn(format="%.2f"),
                }
            )
        
        # Missing in recon
        if recon_results.get('missing_in_recon'):
            st.write(f"**📋 Missing in Recon File ({label})**")
            missing_recon_df = pd.DataFrame(recon_results['missing_in_recon'])
            st.dataframe(missing_recon_df, use_container_width=True, hide_index=True)
        
        # Missing in delivery
        if recon_results.get('missing_in_delivery'):
            st.write(f"**📋 Missing in Delivery Output ({label})**")
            missing_delivery_df = pd.DataFrame(recon_results['missing_in_delivery'])
            st.dataframe(missing_delivery_df, use_container_width=True, hide_index=True)
        
        # If no discrepancies
        if (not recon_results.get('position_mismatches') and 
            not recon_results.get('missing_in_recon') and 
            not recon_results.get('missing_in_delivery')):
            st.success(f"✅ No discrepancies found for {label} positions!")
    
    def generate_consolidated_report(self, delivery_file: str, recon_file: str) -> str:
        """Combine delivery report and reconciliation report into a single Excel file"""
        try:
            from openpyxl import load_workbook
            from copy import copy
            
            # Generate output filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            prefix = getattr(st.session_state, 'file_prefix', 'DELIVERY')
            consolidated_file = f"{prefix}_CONSOLIDATED_{timestamp}.xlsx"
            
            # Load the delivery report as base
            wb_delivery = load_workbook(delivery_file)
            
            # Load the reconciliation report
            wb_recon = load_workbook(recon_file)
            
            # Copy all sheets from reconciliation report to delivery report
            for sheet_name in wb_recon.sheetnames:
                source_sheet = wb_recon[sheet_name]
                
                # Create new sheet name with prefix
                if sheet_name == "Summary":
                    new_sheet_name = "RECON_Summary"
                else:
                    new_sheet_name = f"RECON_{sheet_name}" if not sheet_name.startswith("RECON_") else sheet_name
                
                # Ensure sheet name doesn't exceed Excel's 31 character limit
                if len(new_sheet_name) > 31:
                    new_sheet_name = new_sheet_name[:31]
                
                # Create new sheet in delivery workbook
                target_sheet = wb_delivery.create_sheet(new_sheet_name)
                
                # Copy all cells
                for row in source_sheet.iter_rows():
                    for cell in row:
                        target_cell = target_sheet.cell(
                            row=cell.row, 
                            column=cell.column, 
                            value=cell.value
                        )
                        
                        # Copy cell formatting
                        if cell.has_style:
                            target_cell.font = copy(cell.font)
                            target_cell.fill = copy(cell.fill)
                            target_cell.border = copy(cell.border)
                            target_cell.alignment = copy(cell.alignment)
                            target_cell.number_format = cell.number_format
                
                # Copy column widths
                for col_letter in source_sheet.column_dimensions:
                    target_sheet.column_dimensions[col_letter].width = source_sheet.column_dimensions[col_letter].width
            
            # Save consolidated workbook
            wb_delivery.save(consolidated_file)
            
            logger.info(f"Generated consolidated report: {consolidated_file}")
            return consolidated_file
            
        except Exception as e:
            logger.error(f"Error creating consolidated report: {e}")
            raise
    
    def download_reports_tab(self):
        """Download generated reports"""
        st.markdown('<h2 class="sub-header">Download Reports</h2>', unsafe_allow_html=True)
        
        # Individual reports
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.subheader("📊 Delivery Report")
            
            if not st.session_state.report_generated or not st.session_state.output_file:
                st.info("📤 Please process a position file first")
            else:
                st.success("✅ **Delivery Report Ready!**")
                st.write(f"**Filename:** {st.session_state.output_file}")
                
                if st.session_state.has_trades:
                    st.write("**Contains:**")
                    st.write("- Start, Trade, Final positions")
                    st.write("- Initial & Final calculations")
                
                try:
                    with open(st.session_state.output_file, 'rb') as f:
                        excel_data = f.read()
                    
                    st.download_button(
                        label="📥 Download Delivery Report",
                        data=excel_data,
                        file_name=st.session_state.output_file,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        type="primary"
                    )
                except Exception as e:
                    st.error(f"Error reading report: {str(e)}")
        
        with col2:
            st.subheader("🔄 Reconciliation Report")
            
            if not hasattr(st.session_state, 'recon_output_file'):
                st.info("📋 Upload a recon file and run reconciliation first")
            else:
                st.success("✅ **Reconciliation Report Ready!**")
                st.write(f"**Filename:** {st.session_state.recon_output_file}")
                
                if st.session_state.has_trades:
                    st.write("**Contains:**")
                    st.write("- Initial recon")
                    st.write("- Final recon")
                
                try:
                    with open(st.session_state.recon_output_file, 'rb') as f:
                        recon_data = f.read()
                    
                    st.download_button(
                        label="📥 Download Recon Report",
                        data=recon_data,
                        file_name=st.session_state.recon_output_file,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        type="primary"
                    )
                except Exception as e:
                    st.error(f"Error reading recon report: {str(e)}")
        
        with col3:
            st.subheader("📦 Consolidated Report")
            
            # Check if both reports are available
            if (not st.session_state.report_generated or not st.session_state.output_file or 
                not hasattr(st.session_state, 'recon_output_file')):
                st.info("📋 Generate both reports first")
            else:
                st.success("✅ **Both Reports Available!**")
                st.write("Combine delivery and reconciliation reports into a single file")
                
                if st.button("🔄 Generate Consolidated Report", use_container_width=True, type="primary"):
                    try:
                        with st.spinner("Creating consolidated report..."):
                            consolidated_file = self.generate_consolidated_report(
                                st.session_state.output_file,
                                st.session_state.recon_output_file
                            )
                            
                            # Read the consolidated file for download
                            with open(consolidated_file, 'rb') as f:
                                consolidated_data = f.read()
                            
                            st.download_button(
                                label="📥 Download Consolidated Report",
                                data=consolidated_data,
                                file_name=consolidated_file,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True,
                                type="secondary"
                            )
                            
                            # Clean up the temporary consolidated file
                            try:
                                os.unlink(consolidated_file)
                            except:
                                pass
                                
                    except Exception as e:
                        st.error(f"Error creating consolidated report: {str(e)}")


def main():
    """Main entry point"""
    app = StreamlitDeliveryApp()
    app.run()


if __name__ == "__main__":
    main()
