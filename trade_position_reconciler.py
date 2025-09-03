"""
Trade Position Reconciliation Module - Complete Version
Combines beginning positions with intraday trades to calculate post-trade positions
Generates complete pre-trade and post-trade delivery and IV reports
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple
import logging
from dataclasses import dataclass
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# Import existing modules
from input_parser import Position
from excel_writer import ExcelWriter
from price_fetcher import PriceFetcher

logger = logging.getLogger(__name__)

@dataclass
class TradePosition:
    """Represents a position change from a trade"""
    underlying_ticker: str
    bloomberg_ticker: str
    symbol: str
    expiry_date: datetime
    position_change: float  # Positive for buy, negative for sell
    security_type: str  # Futures, Call, Put
    strike_price: float
    lot_size: int
    trade_price: float
    trade_type: str  # BUY or SELL

class TradePositionReconciler:
    """Main class for reconciling beginning positions with trades"""
    
    def __init__(self, mapping_file: str = "futures mapping.csv", usdinr_rate: float = 88.0):
        self.mapping_file = mapping_file
        self.usdinr_rate = usdinr_rate
        self.symbol_mappings = self._load_mappings()
        
    def _load_mappings(self) -> Dict:
        """Load symbol mappings from CSV"""
        mappings = {}
        try:
            df = pd.read_csv(self.mapping_file)
            for idx, row in df.iterrows():
                if pd.notna(row.iloc[0]) and pd.notna(row.iloc[1]):
                    symbol = str(row.iloc[0]).strip()
                    ticker = str(row.iloc[1]).strip()
                    
                    # Handle underlying
                    underlying = None
                    if len(row) > 2 and pd.notna(row.iloc[2]):
                        underlying_val = str(row.iloc[2]).strip()
                        if underlying_val and underlying_val.upper() != 'NAN':
                            underlying = underlying_val
                    
                    if not underlying:
                        if symbol.upper() in ['NIFTY', 'BANKNIFTY', 'FINNIFTY', 'MIDCPNIFTY']:
                            underlying = f"{symbol.upper()} INDEX"
                        else:
                            underlying = f"{ticker} IS Equity"
                    
                    lot_size = 1
                    if len(row) > 4 and pd.notna(row.iloc[4]):
                        try:
                            lot_size = int(float(str(row.iloc[4]).strip()))
                        except:
                            lot_size = 1
                    
                    mappings[symbol] = {
                        'ticker': ticker,
                        'underlying': underlying,
                        'lot_size': lot_size
                    }
        except Exception as e:
            logger.error(f"Error loading mapping file: {e}")
        
        return mappings
    
    def parse_ms_raw_trade_file(self, trade_file_path: str) -> List[TradePosition]:
        """Parse raw MS format trade file (the input to WAFRA transformer)"""
        trade_positions = []
        
        try:
            # Read the MS input CSV
            df = pd.read_csv(trade_file_path, header=None)
            logger.info(f"Read MS file with {len(df)} rows and {len(df.columns)} columns")
            
            # MS format columns:
            # Col 4: Instrument (OPTSTK, FUTSTK, OPTIDX, FUTIDX)
            # Col 5: Symbol
            # Col 6: Expiry
            # Col 7: Lot Size
            # Col 8: Strike
            # Col 9: Option Type (CE/PE)
            # Col 10: Side (B/S)
            # Col 12: Quantity
            # Col 13: Price
            
            for idx, row in df.iterrows():
                try:
                    # Skip header rows or invalid data
                    if len(row) < 14:
                        continue
                    
                    # Get instrument type
                    instrument = str(row[4]).upper() if pd.notna(row[4]) else ""
                    if instrument not in ['OPTSTK', 'FUTSTK', 'OPTIDX', 'FUTIDX']:
                        continue
                    
                    # Parse fields
                    symbol = str(row[5]).strip().upper() if pd.notna(row[5]) else ""
                    if not symbol:
                        continue
                    
                    expiry_str = str(row[6]).strip() if pd.notna(row[6]) else ""
                    strike = float(row[8]) if pd.notna(row[8]) else 0.0
                    option_type = str(row[9]).strip().upper() if pd.notna(row[9]) else ""
                    side = str(row[10]).strip().upper() if pd.notna(row[10]) else ""
                    qty = float(row[12]) if pd.notna(row[12]) else 0.0
                    price = float(row[13]) if pd.notna(row[13]) else 0.0
                    lot_size = int(float(row[7])) if pd.notna(row[7]) else 1
                    
                    # Skip zero quantity
                    if qty == 0:
                        continue
                    
                    # Parse expiry date
                    expiry = self._parse_date_flexible(expiry_str)
                    if not expiry:
                        logger.warning(f"Could not parse expiry date: {expiry_str}")
                        continue
                    
                    # Determine security type
                    if 'FUT' in instrument:
                        security_type = 'Futures'
                    elif option_type in ['CE', 'C']:
                        security_type = 'Call'
                    elif option_type in ['PE', 'P']:
                        security_type = 'Put'
                    else:
                        continue
                    
                    # Determine position change (Buy = positive, Sell = negative)
                    if side.startswith('B'):
                        position_change = qty
                        trade_type = 'BUY'
                    elif side.startswith('S'):
                        position_change = -qty
                        trade_type = 'SELL'
                    else:
                        continue
                    
                    # Get mapping for underlying and ticker
                    mapping = self.symbol_mappings.get(symbol, {})
                    if not mapping:
                        # Use defaults if not mapped
                        if symbol in ['NIFTY', 'BANKNIFTY', 'FINNIFTY', 'MIDCPNIFTY']:
                            underlying = f"{symbol} INDEX"
                            ticker = symbol
                        else:
                            underlying = f"{symbol} IS Equity"
                            ticker = symbol
                    else:
                        underlying = mapping.get('underlying', f"{symbol} IS Equity")
                        ticker = mapping.get('ticker', symbol)
                    
                    # Generate Bloomberg ticker
                    bloomberg_ticker = self._generate_bloomberg_ticker(
                        ticker, expiry, security_type, strike
                    )
                    
                    # Create TradePosition
                    trade_pos = TradePosition(
                        underlying_ticker=underlying,
                        bloomberg_ticker=bloomberg_ticker,
                        symbol=symbol,
                        expiry_date=expiry,
                        position_change=position_change,
                        security_type=security_type,
                        strike_price=strike,
                        lot_size=lot_size,
                        trade_price=price,
                        trade_type=trade_type
                    )
                    
                    trade_positions.append(trade_pos)
                    logger.debug(f"Added trade: {symbol} {expiry_str} {security_type} {strike} {trade_type} {qty}")
                    
                except Exception as e:
                    logger.debug(f"Error parsing row {idx}: {e}")
                    continue
            
            logger.info(f"Successfully parsed {len(trade_positions)} trades from MS file")
            return trade_positions
            
        except Exception as e:
            logger.error(f"Error reading MS trade file: {e}")
            return []
    
    def parse_wafra_trade_file(self, trade_file_path: str) -> List[TradePosition]:
        """Parse WAFRA format trade file and convert to position changes"""
        trade_positions = []
        
        try:
            # Read the WAFRA output CSV
            df = pd.read_csv(trade_file_path)
            
            for _, row in df.iterrows():
                # Skip rows with UPDATE or missing security
                if pd.isna(row.get('Security')) or row.get('Security') == 'UPDATE':
                    continue
                
                # Parse Bloomberg ticker to extract details
                security = str(row['Security'])
                position_details = self._parse_bloomberg_ticker(security)
                
                if not position_details:
                    continue
                
                # Determine position change
                transaction = str(row.get('Transaction', '')).upper()
                qty = float(row.get('Order_Quantity', 0))
                
                if transaction == 'BUY':
                    position_change = qty
                elif transaction == 'SELL':
                    position_change = -qty
                else:
                    continue
                
                # Get trade price
                trade_price = float(row.get('Order_Price', 0))
                
                # Create TradePosition object
                trade_pos = TradePosition(
                    underlying_ticker=position_details['underlying'],
                    bloomberg_ticker=security,
                    symbol=position_details['symbol'],
                    expiry_date=position_details['expiry'],
                    position_change=position_change,
                    security_type=position_details['type'],
                    strike_price=position_details['strike'],
                    lot_size=int(row.get('Contract Size', 1)),
                    trade_price=trade_price,
                    trade_type=transaction
                )
                
                trade_positions.append(trade_pos)
                
        except Exception as e:
            logger.error(f"Error parsing WAFRA trade file: {e}")
        
        return trade_positions
    
    def _parse_date_flexible(self, date_str: str) -> Optional[datetime]:
        """Parse date string in various formats"""
        date_str = str(date_str).strip()
        
        # Try multiple date formats
        formats = [
            "%d-%b-%Y",  # 27-Mar-2025
            "%d/%m/%Y",  # 27/03/2025
            "%d/%m/%y",  # 27/03/25
            "%Y-%m-%d",  # 2025-03-27
            "%d-%m-%Y",  # 27-03-2025
            "%d.%m.%Y",  # 27.03.2025
            "%d%b%Y",    # 27Mar2025
            "%d-%b-%y",  # 27-Mar-25
            "%d %b %Y",  # 27 Mar 2025
        ]
        
        for fmt in formats:
            try:
                return datetime.strptime(date_str, fmt)
            except:
                continue
        
        # Try pandas parser as fallback
        try:
            return pd.to_datetime(date_str, dayfirst=True)
        except:
            return None
    
    def _parse_bloomberg_ticker(self, ticker: str) -> Optional[Dict]:
        """Parse Bloomberg ticker to extract position details"""
        try:
            ticker = ticker.strip()
            
            # Futures pattern: SYMBOL=MY IS Equity or SYMBOLMY Index
            if '=' in ticker and 'IS Equity' in ticker:
                match = re.match(r'([A-Z]+)=([A-Z])(\d)\s+IS\s+Equity', ticker)
                if match:
                    symbol = match.group(1)
                    month_code = match.group(2)
                    year = match.group(3)
                    
                    month_map = {'F':1,'G':2,'H':3,'J':4,'K':5,'M':6,
                                'N':7,'Q':8,'U':9,'V':10,'X':11,'Z':12}
                    month = month_map.get(month_code, 1)
                    
                    year_full = 2020 + int(year)
                    expiry = self._get_expiry_date(year_full, month)
                    
                    underlying = self._get_underlying_from_symbol(symbol)
                    
                    return {
                        'symbol': symbol,
                        'underlying': underlying,
                        'expiry': expiry,
                        'type': 'Futures',
                        'strike': 0
                    }
            
            # Index futures: NIFTYU5 Index
            elif 'Index' in ticker and '=' not in ticker and not ('/' in ticker):
                match = re.match(r'([A-Z]+)([A-Z])(\d)\s+Index', ticker)
                if match:
                    symbol = match.group(1)
                    month_code = match.group(2)
                    year = match.group(3)
                    
                    month_map = {'F':1,'G':2,'H':3,'J':4,'K':5,'M':6,
                                'N':7,'Q':8,'U':9,'V':10,'X':11,'Z':12}
                    month = month_map.get(month_code, 1)
                    
                    year_full = 2020 + int(year)
                    expiry = self._get_expiry_date(year_full, month)
                    
                    return {
                        'symbol': symbol,
                        'underlying': f"{symbol} INDEX",
                        'expiry': expiry,
                        'type': 'Futures',
                        'strike': 0
                    }
            
            # Options pattern
            elif ('C' in ticker or 'P' in ticker) and '/' in ticker:
                # Stock options
                if 'IS' in ticker and 'Equity' in ticker:
                    match = re.match(r'([A-Z]+)\s+IS\s+(\d{2})/(\d{2})/(\d{2})\s+([CP])(\d+)\s+Equity', ticker)
                    if match:
                        symbol = match.group(1)
                        month = int(match.group(2))
                        day = int(match.group(3))
                        year = 2000 + int(match.group(4))
                        opt_type = 'Call' if match.group(5) == 'C' else 'Put'
                        strike = float(match.group(6))
                        
                        expiry = datetime(year, month, day)
                        underlying = self._get_underlying_from_symbol(symbol)
                        
                        return {
                            'symbol': symbol,
                            'underlying': underlying,
                            'expiry': expiry,
                            'type': opt_type,
                            'strike': strike
                        }
                
                # Index options
                elif 'Index' in ticker:
                    match = re.match(r'([A-Z]+)\s+(\d{2})/(\d{2})/(\d{2})\s+([CP])(\d+)\s+Index', ticker)
                    if match:
                        symbol = match.group(1)
                        month = int(match.group(2))
                        day = int(match.group(3))
                        year = 2000 + int(match.group(4))
                        opt_type = 'Call' if match.group(5) == 'C' else 'Put'
                        strike = float(match.group(6))
                        
                        expiry = datetime(year, month, day)
                        
                        return {
                            'symbol': symbol,
                            'underlying': f"{symbol} INDEX",
                            'expiry': expiry,
                            'type': opt_type,
                            'strike': strike
                        }
            
        except Exception as e:
            logger.debug(f"Could not parse ticker {ticker}: {e}")
        
        return None
    
    def _get_underlying_from_symbol(self, symbol: str) -> str:
        """Get underlying ticker from symbol using mapping"""
        if symbol in self.symbol_mappings:
            return self.symbol_mappings[symbol].get('underlying', f"{symbol} IS Equity")
        return f"{symbol} IS Equity"
    
    def _get_expiry_date(self, year: int, month: int) -> datetime:
        """Calculate expiry date (last Thursday of month typically)"""
        import calendar
        
        # Get last day of month
        last_day = calendar.monthrange(year, month)[1]
        
        # Find last Thursday
        for day in range(last_day, 0, -1):
            date_obj = datetime(year, month, day)
            if date_obj.weekday() == 3:  # Thursday
                return date_obj
        
        return datetime(year, month, last_day)
    
    def _generate_bloomberg_ticker(self, ticker: str, expiry: datetime,
                                  security_type: str, strike: float) -> str:
        """Generate Bloomberg ticker format"""
        
        # Month codes for futures
        MONTH_CODE = {
            1: "F", 2: "G", 3: "H", 4: "J", 5: "K", 6: "M",
            7: "N", 8: "Q", 9: "U", 10: "V", 11: "X", 12: "Z"
        }
        
        ticker_upper = ticker.upper()
        
        # Check if this is an index
        is_index = ticker_upper in ['NIFTY', 'BANKNIFTY', 'FINNIFTY', 'MIDCPNIFTY', 'NZ', 'NBZ', 'NF', 'NBF']
        
        if security_type == 'Futures':
            month_code = MONTH_CODE.get(expiry.month, "")
            year_code = str(expiry.year)[-1]
            
            if is_index:
                return f"{ticker}{month_code}{year_code} Index"
            else:
                return f"{ticker}={month_code}{year_code} IS Equity"
        else:
            # Options
            date_str = expiry.strftime('%m/%d/%y')
            strike_str = str(int(strike)) if strike == int(strike) else str(strike)
            
            if is_index:
                if security_type == 'Call':
                    return f"{ticker} {date_str} C{strike_str} Index"
                else:
                    return f"{ticker} {date_str} P{strike_str} Index"
            else:
                if security_type == 'Call':
                    return f"{ticker} IS {date_str} C{strike_str} Equity"
                else:
                    return f"{ticker} IS {date_str} P{strike_str} Equity"
    
    def read_beginning_positions(self, excel_file: str) -> pd.DataFrame:
        """Read All_Positions sheet from beginning positions file"""
        try:
            df = pd.read_excel(excel_file, sheet_name='All_Positions')
            return df
        except Exception as e:
            logger.error(f"Error reading beginning positions: {e}")
            return pd.DataFrame()
    
    def combine_positions(self, beginning_df: pd.DataFrame, 
                         trade_positions: List[TradePosition]) -> pd.DataFrame:
        """Combine beginning positions with trades to get post-trade positions"""
        
        # Convert beginning positions to dictionary for easier lookup
        position_dict = {}
        
        for _, row in beginning_df.iterrows():
            key = (
                str(row['Symbol']),
                pd.to_datetime(row['Expiry']).strftime('%Y-%m-%d'),
                str(row['Type']),
                float(row['Strike']) if pd.notna(row['Strike']) else 0
            )
            
            position_dict[key] = {
                'underlying': row['Underlying'],
                'symbol': row['Symbol'],
                'expiry': pd.to_datetime(row['Expiry']),
                'position': float(row['Position']),
                'type': row['Type'],
                'strike': float(row['Strike']) if pd.notna(row['Strike']) else 0,
                'lot_size': int(row['Lot Size']) if pd.notna(row['Lot Size']) else 1
            }
        
        # Apply trades
        for trade in trade_positions:
            key = (
                trade.bloomberg_ticker,
                trade.expiry_date.strftime('%Y-%m-%d'),
                trade.security_type,
                trade.strike_price
            )
            
            if key in position_dict:
                # Update existing position
                position_dict[key]['position'] += trade.position_change
            else:
                # New position from trade
                position_dict[key] = {
                    'underlying': trade.underlying_ticker,
                    'symbol': trade.bloomberg_ticker,
                    'expiry': trade.expiry_date,
                    'position': trade.position_change,
                    'type': trade.security_type,
                    'strike': trade.strike_price,
                    'lot_size': trade.lot_size
                }
        
        # Convert back to DataFrame
        post_trade_data = []
        for key, pos in position_dict.items():
            # Only include non-zero positions
            if pos['position'] != 0:
                post_trade_data.append({
                    'Underlying': pos['underlying'],
                    'Symbol': pos['symbol'],
                    'Expiry': pos['expiry'].strftime('%Y-%m-%d'),
                    'Position': pos['position'],
                    'Type': pos['type'],
                    'Strike': pos['strike'] if pos['strike'] > 0 else '',
                    'Lot Size': pos['lot_size']
                })
        
        post_trade_df = pd.DataFrame(post_trade_data)
        
        # Sort by underlying, expiry, strike
        if not post_trade_df.empty:
            post_trade_df = post_trade_df.sort_values(
                by=['Underlying', 'Expiry', 'Strike'],
                ascending=[True, True, True]
            )
        
        return post_trade_df
    
    def generate_post_trade_report(self, beginning_file: str, trade_file: str, 
                                  output_file: str, trade_format: str = 'MS'):
        """Main method to generate complete post-trade report"""
        
        logger.info("Starting post-trade position reconciliation...")
        
        # Step 1: Read beginning positions
        logger.info("Reading beginning positions...")
        beginning_df = self.read_beginning_positions(beginning_file)
        
        if beginning_df.empty:
            logger.error("No beginning positions found")
            return False
        
        logger.info(f"Found {len(beginning_df)} beginning positions")
        
        # Step 2: Parse trade file
        logger.info(f"Parsing {trade_format} trade file...")
        
        if trade_format == 'MS':
            trade_positions = self.parse_ms_raw_trade_file(trade_file)
        elif trade_format == 'WAFRA':
            trade_positions = self.parse_wafra_trade_file(trade_file)
        else:
            logger.error(f"Unsupported trade format: {trade_format}")
            return False
        
        logger.info(f"Parsed {len(trade_positions)} trades")
        
        # Step 3: Combine positions
        logger.info("Combining positions...")
        post_trade_df = self.combine_positions(beginning_df, trade_positions)
        
        logger.info(f"Post-trade positions: {len(post_trade_df)}")
        
        # Step 4: Convert to Position objects for report generation
        positions = []
        for _, row in post_trade_df.iterrows():
            pos = Position(
                underlying_ticker=row['Underlying'],
                bloomberg_ticker=row['Symbol'],
                symbol=row['Symbol'].split()[0],
                expiry_date=pd.to_datetime(row['Expiry']),
                position_lots=row['Position'],
                security_type=row['Type'],
                strike_price=float(row['Strike']) if row['Strike'] else 0,
                lot_size=row['Lot Size']
            )
            positions.append(pos)
        
        # Step 5: Fetch prices
        logger.info("Fetching current prices...")
        price_fetcher = PriceFetcher()
        symbols_to_fetch = list(set(p.symbol for p in positions))
        prices = price_fetcher.fetch_prices_for_symbols(symbols_to_fetch)
        
        # Step 6: Create enhanced Excel report
        logger.info("Creating post-trade Excel report...")
        self._create_enhanced_report(
            output_file, beginning_df, trade_positions, 
            post_trade_df, positions, prices
        )
        
        logger.info(f"Post-trade report saved: {output_file}")
        return True
    
    def _create_enhanced_report(self, output_file: str, beginning_df: pd.DataFrame,
                               trade_positions: List[TradePosition], 
                               post_trade_df: pd.DataFrame,
                               positions: List[Position], 
                               prices: Dict[str, float]):
        """Create Excel report with complete pre and post trade positions"""
        
        # Create the Excel writer
        writer = ExcelWriter(output_file, self.usdinr_rate)
        wb = writer.wb
        
        # Define styles
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        summary_header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        trade_header_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # ============= 1. SUMMARY SHEET =============
        ws_summary = wb.create_sheet("Summary", 0)
        ws_summary.cell(1, 1, "TRADE RECONCILIATION SUMMARY").font = Font(bold=True, size=14)
        ws_summary.cell(3, 1, "Report Generated:").font = Font(bold=True)
        ws_summary.cell(3, 2, datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        
        # Summary statistics
        ws_summary.cell(5, 1, "POSITION SUMMARY").font = Font(bold=True, size=12)
        ws_summary.cell(5, 1).fill = summary_header_fill
        
        summary_data = [
            ("Beginning Positions Count", len(beginning_df)),
            ("Trades Executed", len(trade_positions)),
            ("Post-Trade Positions Count", len(post_trade_df)),
            ("Net New Positions", len(post_trade_df) - len(beginning_df))
        ]
        
        row = 6
        for label, value in summary_data:
            ws_summary.cell(row, 1, label).font = Font(bold=True)
            ws_summary.cell(row, 2, value)
            row += 1
        
        # Position changes by underlying
        ws_summary.cell(row + 1, 1, "POSITION CHANGES BY UNDERLYING").font = Font(bold=True, size=12)
        ws_summary.cell(row + 1, 1).fill = summary_header_fill
        row += 2
        
        # Headers for changes table
        change_headers = ["Underlying", "Pre-Trade", "Trades", "Post-Trade", "Net Change"]
        for col, header in enumerate(change_headers, 1):
            cell = ws_summary.cell(row, col, header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        row += 1
        
        # Calculate changes
        begin_by_underlying = beginning_df.groupby('Underlying')['Position'].sum()
        post_by_underlying = post_trade_df.groupby('Underlying')['Position'].sum()
        
        # Calculate trade quantities by underlying
        trade_by_underlying = {}
        for trade in trade_positions:
            underlying = trade.underlying_ticker
            if underlying not in trade_by_underlying:
                trade_by_underlying[underlying] = 0
            trade_by_underlying[underlying] += trade.position_change
        
        all_underlyings = set(list(begin_by_underlying.index) + 
                              list(post_by_underlying.index) + 
                              list(trade_by_underlying.keys()))
        
        for underlying in sorted(all_underlyings):
            begin = begin_by_underlying.get(underlying, 0)
            trades = trade_by_underlying.get(underlying, 0)
            post = post_by_underlying.get(underlying, 0)
            change = post - begin
            
            ws_summary.cell(row, 1, underlying).border = border
            ws_summary.cell(row, 2, begin).border = border
            ws_summary.cell(row, 3, trades).border = border
            ws_summary.cell(row, 4, post).border = border
            
            change_cell = ws_summary.cell(row, 5, change)
            change_cell.border = border
            if change > 0:
                change_cell.font = Font(color="008000", bold=True)  # Green
            elif change < 0:
                change_cell.font = Font(color="FF0000", bold=True)  # Red
            
            row += 1
        
        # Set column widths
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 15
        ws_summary.column_dimensions['C'].width = 15
        ws_summary.column_dimensions['D'].width = 15
        ws_summary.column_dimensions['E'].width = 15
        
        # ============= 2. TRADES SHEET =============
        ws_trades = wb.create_sheet("Trades", 1)
        ws_trades.cell(1, 1, "INTRADAY TRADES").font = Font(bold=True, size=12)
        
        trade_headers = ['Trade #', 'Underlying', 'Symbol', 'Bloomberg Ticker', 'Expiry', 
                        'Type', 'Strike', 'Side', 'Quantity', 'Price', 'Lot Size']
        
        for col, header in enumerate(trade_headers, 1):
            cell = ws_trades.cell(3, col, header)
            cell.font = header_font
            cell.fill = trade_header_fill
            cell.border = border
        
        row = 4
        for idx, trade in enumerate(trade_positions, 1):
            ws_trades.cell(row, 1, idx).border = border
            ws_trades.cell(row, 2, trade.underlying_ticker).border = border
            ws_trades.cell(row, 3, trade.symbol).border = border
            ws_trades.cell(row, 4, trade.bloomberg_ticker).border = border
            ws_trades.cell(row, 5, trade.expiry_date.strftime('%Y-%m-%d')).border = border
            ws_trades.cell(row, 6, trade.security_type).border = border
            ws_trades.cell(row, 7, trade.strike_price if trade.strike_price > 0 else '').border = border
            ws_trades.cell(row, 8, trade.trade_type).border = border
            ws_trades.cell(row, 9, abs(trade.position_change)).border = border
            ws_trades.cell(row, 10, trade.trade_price).border = border
            ws_trades.cell(row, 11, trade.lot_size).border = border
            row += 1
        
        # Set column widths for trades sheet
        ws_trades.column_dimensions['A'].width = 10
        ws_trades.column_dimensions['B'].width = 25
        ws_trades.column_dimensions['C'].width = 15
        ws_trades.column_dimensions['D'].width = 30
        ws_trades.column_dimensions['E'].width = 12
        ws_trades.column_dimensions['F'].width = 10
        ws_trades.column_dimensions['G'].width = 10
        ws_trades.column_dimensions['H'].width = 10
        ws_trades.column_dimensions['I'].width = 12
        ws_trades.column_dimensions['J'].width = 10
        ws_trades.column_dimensions['K'].width = 10
        
        # ============= 3. PRE-TRADE POSITIONS =============
        # Convert beginning positions to Position objects
        pre_positions = []
        for _, row in beginning_df.iterrows():
            pos = Position(
                underlying_ticker=row['Underlying'],
                bloomberg_ticker=row['Symbol'],
                symbol=row['Symbol'].split()[0] if ' ' in row['Symbol'] else row['Symbol'],
                expiry_date=pd.to_datetime(row['Expiry']),
                position_lots=float(row['Position']),
                security_type=row['Type'],
                strike_price=float(row['Strike']) if row.get('Strike') and str(row['Strike']).strip() else 0,
                lot_size=int(row['Lot Size']) if row.get('Lot Size') else 1
            )
            pre_positions.append(pos)
        
        # 3a. Pre-Trade All Positions
        ws_pre_all = wb.create_sheet("PreTrade_All_Positions")
        self._write_positions_sheet(ws_pre_all, beginning_df, header_font, header_fill, border)
        
        # 3b. Pre-Trade Master Deliverables
        ws_pre_master = wb.create_sheet("PreTrade_Master_All_Expiries")
        self._write_deliverable_sheet(ws_pre_master, pre_positions, prices, "PRE-TRADE", writer)
        
        # 3c. Pre-Trade Expiry-wise Deliverables
        pre_expiries = list(set(p.expiry_date for p in pre_positions))
        for expiry in sorted(pre_expiries):
            sheet_name = f"PreTrade_Expiry_{expiry.strftime('%Y_%m_%d')}"
            ws_pre_exp = wb.create_sheet(sheet_name)
            expiry_positions = [p for p in pre_positions if p.expiry_date.date() == expiry.date()]
            self._write_deliverable_sheet(ws_pre_exp, expiry_positions, prices, f"PRE-TRADE {expiry.strftime('%Y-%m-%d')}", writer)
        
        # 3d. Pre-Trade IV Master
        ws_pre_iv_master = wb.create_sheet("PreTrade_IV_All_Expiries")
        self._write_iv_sheet(ws_pre_iv_master, pre_positions, prices, "PRE-TRADE", writer)
        
        # 3e. Pre-Trade IV Expiry-wise
        for expiry in sorted(pre_expiries):
            sheet_name = f"PreTrade_IV_{expiry.strftime('%Y_%m_%d')}"
            ws_pre_iv_exp = wb.create_sheet(sheet_name)
            expiry_positions = [p for p in pre_positions if p.expiry_date.date() == expiry.date()]
            self._write_iv_sheet(ws_pre_iv_exp, expiry_positions, prices, f"PRE-TRADE IV {expiry.strftime('%Y-%m-%d')}", writer)
        
        # ============= 4. POST-TRADE POSITIONS =============
        # 4a. Post-Trade All Positions
        ws_post_all = wb.create_sheet("PostTrade_All_Positions")
        self._write_positions_sheet(ws_post_all, post_trade_df, header_font, header_fill, border)
        
        # 4b. Post-Trade Master Deliverables
        ws_post_master = wb.create_sheet("PostTrade_Master_All_Expiries")
        self._write_deliverable_sheet(ws_post_master, positions, prices, "POST-TRADE", writer)
        
        # 4c. Post-Trade Expiry-wise Deliverables
        post_expiries = list(set(p.expiry_date for p in positions))
        for expiry in sorted(post_expiries):
            sheet_name = f"PostTrade_Expiry_{expiry.strftime('%Y_%m_%d')}"
            ws_post_exp = wb.create_sheet(sheet_name)
            expiry_positions = [p for p in positions if p.expiry_date.date() == expiry.date()]
            self._write_deliverable_sheet(ws_post_exp, expiry_positions, prices, f"POST-TRADE {expiry.strftime('%Y-%m-%d')}", writer)
        
        # 4d. Post-Trade IV Master
        ws_post_iv_master = wb.create_sheet("PostTrade_IV_All_Expiries")
        self._write_iv_sheet(ws_post_iv_master, positions, prices, "POST-TRADE", writer)
        
        # 4e. Post-Trade IV Expiry-wise
        for expiry in sorted(post_expiries):
            sheet_name = f"PostTrade_IV_{expiry.strftime('%Y_%m_%d')}"
            ws_post_iv_exp = wb.create_sheet(sheet_name)
            expiry_positions = [p for p in positions if p.expiry_date.date() == expiry.date()]
            self._write_iv_sheet(ws_post_iv_exp, expiry_positions, prices, f"POST-TRADE IV {expiry.strftime('%Y-%m-%d')}", writer)
        
        # Save the workbook
        wb.save(output_file)
        logger.info(f"Enhanced report saved with pre-trade and post-trade sheets: {output_file}")
    
    def _write_deliverable_sheet(self, ws, positions: List[Position], prices: Dict[str, float], 
                                title: str, writer: ExcelWriter):
        """Write deliverable sheet using ExcelWriter logic"""
        
        # Title
        ws.cell(1, 1, f"{title} DELIVERABLES").font = Font(bold=True, size=12)
        
        # Headers
        headers = [
            "Underlying", "Symbol", "Expiry", "Position", "Type", "Strike",
            "System Deliverable", "Override Deliverable", "System Price",
            "Override Price", "BBG Price", "BBG Deliverable"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(3, col, header)
            cell.font = writer.header_font
            cell.fill = writer.header_fill
            cell.alignment = writer.header_alignment
            cell.border = writer.border
        
        # Group positions by underlying
        grouped = {}
        for pos in positions:
            if pos.underlying_ticker not in grouped:
                grouped[pos.underlying_ticker] = []
            grouped[pos.underlying_ticker].append(pos)
        
        current_row = 4
        
        for underlying in sorted(grouped.keys()):
            underlying_positions = grouped[underlying]
            group_start_row = current_row
            
            # Group header
            cell = ws.cell(current_row, 1, underlying)
            cell.font = writer.group_font
            
            for col in range(1, 13):
                cell = ws.cell(current_row, col)
                cell.fill = writer.group_fill
                cell.border = writer.border
            
            # Get price
            symbol = underlying_positions[0].symbol
            spot_price = prices.get(symbol)
            
            if spot_price:
                price_cell = ws.cell(current_row, 9, spot_price)
                price_cell.number_format = writer.price_format
            
            current_row += 1
            detail_rows = []
            
            # Sort and write detail rows
            underlying_positions.sort(key=lambda x: (x.expiry_date, x.strike_price))
            
            for pos in underlying_positions:
                ws.cell(current_row, 2, pos.bloomberg_ticker)
                ws.cell(current_row, 3, pos.expiry_date.strftime('%Y-%m-%d'))
                ws.cell(current_row, 4, pos.position_lots)
                ws.cell(current_row, 5, pos.security_type)
                
                if pos.strike_price > 0:
                    strike_cell = ws.cell(current_row, 6, pos.strike_price)
                    strike_cell.number_format = writer.price_format
                
                # Deliverable calculation
                if spot_price:
                    if pos.security_type == 'Futures':
                        deliverable = pos.position_lots
                    elif pos.security_type == 'Call':
                        deliverable = pos.position_lots if spot_price > pos.strike_price else 0
                    elif pos.security_type == 'Put':
                        deliverable = -pos.position_lots if spot_price < pos.strike_price else 0
                    else:
                        deliverable = 0
                    
                    deliv_cell = ws.cell(current_row, 7, deliverable)
                    deliv_cell.number_format = writer.deliv_format
                
                for col in range(1, 13):
                    ws.cell(current_row, col).border = writer.border
                
                detail_rows.append(current_row)
                current_row += 1
            
            # Group totals
            if detail_rows and spot_price:
                total = sum(ws.cell(row, 7).value or 0 for row in detail_rows)
                total_cell = ws.cell(group_start_row, 7, total)
                total_cell.number_format = writer.deliv_format
        
        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 12
        ws.column_dimensions['K'].width = 12
        ws.column_dimensions['L'].width = 15
    
    def _write_iv_sheet(self, ws, positions: List[Position], prices: Dict[str, float], 
                       title: str, writer: ExcelWriter):
        """Write IV sheet using ExcelWriter logic"""
        
        # Title and grand total row
        ws.cell(1, 1, f"{title} INTRINSIC VALUE").font = Font(bold=True, size=12)
        
        grand_total_row = 2
        ws.cell(grand_total_row, 1, "GRAND TOTAL").font = Font(bold=True, size=12)
        
        for col in range(1, 11):
            cell = ws.cell(grand_total_row, col)
            cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            cell.border = writer.border
        
        # Headers
        headers = [
            "Underlying", "Symbol", "Expiry", "Position", "Type", "Strike", "Lot Size",
            "IV (INR)", "IV (USD)", "Price"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(4, col, header)
            cell.font = writer.header_font
            cell.fill = writer.header_fill
            cell.alignment = writer.header_alignment
            cell.border = writer.border
        
        # Group positions by underlying
        grouped = {}
        for pos in positions:
            if pos.underlying_ticker not in grouped:
                grouped[pos.underlying_ticker] = []
            grouped[pos.underlying_ticker].append(pos)
        
        current_row = 5
        total_iv_inr = 0
        total_iv_usd = 0
        
        for underlying in sorted(grouped.keys()):
            underlying_positions = grouped[underlying]
            
            # Group header
            cell = ws.cell(current_row, 1, underlying)
            cell.font = writer.group_font
            
            for col in range(1, 11):
                cell = ws.cell(current_row, col)
                cell.fill = writer.group_fill
                cell.border = writer.border
            
            # Get price
            symbol = underlying_positions[0].symbol
            spot_price = prices.get(symbol, 0)
            
            if spot_price:
                price_cell = ws.cell(current_row, 10, spot_price)
                price_cell.number_format = writer.price_format
            
            current_row += 1
            group_iv_inr = 0
            
            # Sort and write detail rows
            underlying_positions.sort(key=lambda x: (x.expiry_date, x.strike_price))
            
            for pos in underlying_positions:
                ws.cell(current_row, 2, pos.bloomberg_ticker)
                ws.cell(current_row, 3, pos.expiry_date.strftime('%Y-%m-%d'))
                ws.cell(current_row, 4, pos.position_lots)
                ws.cell(current_row, 5, pos.security_type)
                
                if pos.strike_price > 0:
                    strike_cell = ws.cell(current_row, 6, pos.strike_price)
                    strike_cell.number_format = writer.price_format
                
                ws.cell(current_row, 7, pos.lot_size)
                
                # IV calculation
                iv_inr = 0
                if spot_price:
                    if pos.security_type == 'Call' and spot_price > pos.strike_price:
                        iv_inr = pos.position_lots * pos.lot_size * (spot_price - pos.strike_price)
                    elif pos.security_type == 'Put' and spot_price < pos.strike_price:
                        iv_inr = pos.position_lots * pos.lot_size * (pos.strike_price - spot_price)
                
                iv_usd = iv_inr / writer.usdinr_rate
                
                iv_inr_cell = ws.cell(current_row, 8, iv_inr)
                iv_inr_cell.number_format = writer.iv_format
                
                iv_usd_cell = ws.cell(current_row, 9, iv_usd)
                iv_usd_cell.number_format = writer.iv_format
                
                group_iv_inr += iv_inr
                
                for col in range(1, 11):
                    ws.cell(current_row, col).border = writer.border
                
                current_row += 1
            
            # Group total
            group_iv_usd = group_iv_inr / writer.usdinr_rate
            total_iv_inr += group_iv_inr
            total_iv_usd += group_iv_usd
        
        # Grand totals
        ws.cell(grand_total_row, 8, total_iv_inr).number_format = writer.iv_format
        ws.cell(grand_total_row, 9, total_iv_usd).number_format = writer.iv_format
        
        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 12
    
    def _write_positions_sheet(self, ws, df, header_font, header_fill, border):
        """Helper to write positions data to worksheet"""
        headers = ['Underlying', 'Symbol', 'Expiry', 'Position', 'Type', 'Strike', 'Lot Size']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        for row_idx, (_, row) in enumerate(df.iterrows(), 2):
            for col_idx, header in enumerate(headers, 1):
                value = row.get(header, '')
                cell = ws.cell(row_idx, col_idx, value)
                cell.border = border
        
        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 10
    
    def _write_trades_sheet(self, ws, trades, header_font, header_fill, border):
        """Helper to write trades data to worksheet"""
        headers = ['Underlying', 'Symbol', 'Expiry', 'Type', 'Strike', 
                  'Trade Type', 'Quantity', 'Price', 'Position Change']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        for row_idx, trade in enumerate(trades, 2):
            ws.cell(row_idx, 1, trade.underlying_ticker).border = border
            ws.cell(row_idx, 2, trade.bloomberg_ticker).border = border
            ws.cell(row_idx, 3, trade.expiry_date.strftime('%Y-%m-%d')).border = border
            ws.cell(row_idx, 4, trade.security_type).border = border
            ws.cell(row_idx, 5, trade.strike_price if trade.strike_price > 0 else '').border = border
            ws.cell(row_idx, 6, trade.trade_type).border = border
            ws.cell(row_idx, 7, abs(trade.position_change)).border = border
            ws.cell(row_idx, 8, trade.trade_price).border = border
            ws.cell(row_idx, 9, trade.position_change).border = border
        
        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 15
